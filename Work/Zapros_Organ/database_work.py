import cx_Oracle
import pyodbc
import pandas as pd
import openpyxl
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
import os

dsn_tns = cx_Oracle.makedsn('DPE', '1521', service_name='ORCL')


# conn_ora = cx_Oracle.connect(user='LPU', password='Ring43', dsn=dsn_tns)
# conn_ms = pyodbc.connect("""Driver={SQL Server Native Client 11.0};Server=dpx\mssqlserver2012;
#                 Database=my_base;uid=srz_admin;pwd=srz_admin""")

class Oracle(object):

    def connect(self, username, password, hostname, port, servicename):
        try:
            self.db = cx_Oracle.connect(username, password, hostname + ':' + port + '/' + servicename)
        except cx_Oracle.DatabaseError as e:
            raise
        self.cursor = self.db.cursor()

    def disconnect(self):
        try:
            self.cursor.close()
            self.db.close()
        except cx_Oracle.DatabaseError:
            pass

    def execute(self, sql, bindvars=None, commit=False):
        try:
            self.cursor.execute(sql, bindvars)
        except cx_Oracle.DatabaseError as e:
            raise
        # Only commit if it-s necessary.
        if commit:
            self.db.commit()


# if __name__ == "__main__":
#
#     oracle = Oracle.connect('username', 'password', 'hostname'
#                            , 'port', 'servicename')
#
#     try:
#         # No commit as you don-t need to commit DDL.
#         oracle.execute('ddl_statements')
#
#     # Ensure that we always disconnect from the database to avoid
#     # ORA-00018: Maximum number of sessions exceeded.
#     finally:
#         oracle.disconnect()

def make_zapros(**args):
    # params = {"fam": 'Львов', "im": 'Максим', "ot": 'Александрович'}
    params = {}
    for x, value in args.items():
        if value:
            params[x] = value
    print(params)

    sql_ora = '''
        select FIO, to_char(dr, 'dd.mm.yyyy') DR, POLIS, SMO, PROF, MKB, MKB_NAME, PERIOD, SUMV, SUMP, MO, USL_OK,
                DT1, OGRN, ADDRES, TEL, DOCNAME, DOC
        from LPU.VW_LPU_ZAPROS_ORGAN t where rownum <= 1000 and'''
    add_sql = ''
    for key, value in params.items():
        if key == 'fam':
            add_sql += f" lower({key}) = lower(:{key}) and"
        elif key in ('im', 'ot'):
            add_sql += f" lower({key}) LIKE lower(:{key}||'%') and"
        elif key == 'date_begin':
            add_sql += f' dt2 >= :{key} and'
        elif key == 'date_end':
            add_sql += f' dt2 <= :{key} and'
        else:
            add_sql += f' {key} = :{key} and'
    add_sql = add_sql[:-4]
    add_sql += ' ORDER BY dt2'
    sql_ora += add_sql

    sql_ms = '''
        SELECT top 1000 FIO, CONVERT(VARCHAR, dr, 104) DR, POLIS, SMO, PROF, MKB, MKB_NAME, PERIOD, SUMV, SUMP, MO, USL_OK,
        DT1, OGRN, ADDRES, TEL, DOCNAME, DOC
        FROM [IESDB].[dbo].[K_V_Zapros_Organ] where'''
    add_sql = ''
    for key, value in params.items():
        if key in ('im', 'ot'):
            add_sql += f" {key} LIKE ?+'%' and"
        elif key == 'date_begin':
            add_sql += f' dt2 >= ? and'
        elif key == 'date_end':
            add_sql += f' dt2 <= ? and'
        else:
            add_sql += f' {key} = ? and'
    add_sql = add_sql[:-4]
    add_sql += ' ORDER BY dt2'
    sql_ms += add_sql

    with cx_Oracle.connect(user='LPU', password='Ring43', dsn=dsn_tns) as conn_ora:

        try:
            df_ora = pd.read_sql(sql_ora, con=conn_ora, params=params)
        except cx_Oracle.DatabaseError as e:
            return pd.DataFrame(), 'Ошибка подключения к Oracle - ' + e.args[0].message

    with pyodbc.connect("""Driver={SQL Server Native Client 11.0};Server=dpx\mssqlserver2012;
#                 Database=my_base;uid=srz_admin;pwd=srz_admin""") as conn_ms:
        conn_ms.timeout = 30
        try:
            df_ms = pd.read_sql(sql_ms, con=conn_ms, params=params.values(), )
        except pyodbc.Error as ex:
            return pd.DataFrame(), 'Ошибка подключения к MS Sql Server - ' + ex.args[0].message

    df = df_ora.append(df_ms)
    df.columns = ['Фамилия Имя Отчество', 'Дата рождения', 'Серия, номер полиса',
                  'Наименование СМО', 'Профиль/специальность', 'МКБ', 'Диагноз', 'Период лечения', 'Сумма выставленная',
                  'Сумма принятая', 'Наименование МО', 'Условия оказания МП', 'Дата начала лечения',
                  'ОГРН МО', 'Адрес МО', 'Телефон МО', 'Тип документа', 'Серия и номер документа']

    if len(df.index) >= 1000:
        comment = 'Записей больше 1000. Уточните запрос'
        return df, comment
    if not df.empty:
        comment = 'Результаты поиска'
    else:
        comment = 'Нет данных'
    return df, comment


def save_excel(df):
    print('Save Excel')
    wb = openpyxl.load_workbook('zapros_template.xlsx')
    ws = wb.active
    new_df = df.drop(
        ['Дата начала лечения', 'ОГРН МО', 'Адрес МО', 'Телефон МО', 'Тип документа', 'Серия и номер документа'],
        axis=1)

    rows = dataframe_to_rows(new_df, index=False, header=None)

    for r_idx, row in enumerate(rows, 8):
        ws.insert_rows(r_idx + 1, 1)
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx)._style = copy(ws.cell(row=r_idx - 1, column=c_idx)._style)
    wb.save('temp.xlsx')
    # os.startfile(r'temp.xlsx')


def save_excel_2_FSS(df):
    print('Save Excel 2 FSS')
    wb = openpyxl.load_workbook('zapros_2_template.xlsx')
    ws = wb.active
    new_df = df[
        ['Фамилия Имя Отчество', 'Серия, номер полиса', 'Дата рождения', 'Тип документа', 'Серия и номер документа',
         'Дата начала лечения', 'МКБ',
         'Наименование МО', 'ОГРН МО', 'Адрес МО', 'Телефон МО'
         ]].copy()
    new_df.insert(5, 'UDL_ORGAN', '')
    new_df.insert(6, 'UDL_DATE', '')
    new_df.insert(7, 'D_SLUCH', '')
    rows = dataframe_to_rows(new_df, index=False, header=None)

    for r_idx, row in enumerate(rows, 13):
        ws.insert_rows(r_idx + 1, 1)
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx)._style = copy(ws.cell(row=r_idx - 1, column=c_idx)._style)
    wb.save('temp_2_FSS.xlsx')

# d = make_zapros(fam='Львов', im='Максим', ot='Александрович')
# print(d)
